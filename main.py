"""
RAG System - Main FastAPI Server
=================================
This is the entry point for the RAG application.

It provides:
- REST API endpoints for uploading documents and asking questions
- Static file serving for the web interface
- Health check endpoint

Run with: python main.py
Then open: http://localhost:8000
"""

import os
import asyncio
import shutil
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from document_processor import DocumentProcessor
from vector_store import VectorStore
from rag_engine import RAGEngine
from admin_routes import router as admin_router, init_admin
from web_search_service import WebSearchService
from query_analyzer import QueryAnalyzer, SearchIntent

# Create necessary directories
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

STATIC_DIR = Path("static")
STATIC_DIR.mkdir(exist_ok=True)

# Initialize components
# These are created once when the server starts
print("=" * 50)
print("Initializing RAG System...")
print("=" * 50)

document_processor = DocumentProcessor(chunk_size=350, chunk_overlap=35)
vector_store = VectorStore(persist_directory="chroma_data")
rag_engine = RAGEngine(vector_store)
web_search = WebSearchService()
query_analyzer = QueryAnalyzer()

print("=" * 50)
print("RAG System Ready!")
print("=" * 50)

# Initialize admin routes with vector store and upload directory
init_admin(vector_store, UPLOAD_DIR)

# Create FastAPI app
app = FastAPI(
    title="USF RAG System",
    description="A simple RAG chatbot that answers questions about your documents",
    version="1.0.0"
)

# Allow CORS (needed if frontend is served separately during development)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include admin routes
app.include_router(admin_router)


# ============================================================================
# Pydantic Models (for request/response validation)
# ============================================================================

class ChatRequest(BaseModel):
    """Request body for asking a question."""
    question: str
    n_chunks: Optional[int] = 5  # How many context chunks to use (default 5 for better context)
    filter_sources: Optional[List[str]] = None  # Filter search to specific documents
    force_web_search: Optional[bool] = False  # Force web search mode


class ChatResponse(BaseModel):
    """Response body for a question answer."""
    answer: str
    sources: List[str]
    chunks_used: List[dict]


class DocumentInfo(BaseModel):
    """Information about an uploaded document."""
    filename: str
    chunks: int


class FileUploadResult(BaseModel):
    """Result of a single file upload."""
    filename: str
    status: str  # 'success' or 'error'
    chunks: Optional[int] = None
    error: Optional[str] = None


class MultiUploadResponse(BaseModel):
    """Response for multi-file upload."""
    total_files: int
    successful: int
    failed: int
    results: List[FileUploadResult]


class StatsResponse(BaseModel):
    """System statistics."""
    total_chunks: int
    documents: List[str]


# ============================================================================
# API Endpoints
# ============================================================================

@app.get("/")
async def root():
    """Serve the main web interface."""
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/login")
async def login_page():
    """Serve the login page."""
    return FileResponse(STATIC_DIR / "login.html")


@app.get("/health")
async def health_check():
    """Health check endpoint - useful for monitoring."""
    return {"status": "healthy", "message": "RAG system is running!"}


async def process_single_file(file: UploadFile) -> FileUploadResult:
    """
    Process a single file upload. Used by both single and multi-upload endpoints.
    Returns FileUploadResult with status.
    
    For images: Uses quick local processing first, then enhances with Vision API in background.
    For other files: Processes normally (fast for text/PDF).
    """
    filename = file.filename
    
    # Read file content to get size
    content = await file.read()
    file_size = len(content)
    
    # Validate file
    validation = DocumentProcessor.validate_file(filename, file_size)
    if not validation['valid']:
        return FileUploadResult(
            filename=filename,
            status='error',
            error=validation['error']
        )
    
    # Save the file
    file_path = UPLOAD_DIR / filename
    try:
        with open(file_path, "wb") as buffer:
            buffer.write(content)
    except Exception as e:
        return FileUploadResult(
            filename=filename,
            status='error',
            error=f"Failed to save file: {str(e)}"
        )
    
    # Check if it's an image - use quick indexing first
    ext = Path(filename).suffix.lower()
    is_image = ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']
    
    # Process the document - all processing happens in background for instant uploads
    try:
        if is_image:
            # For images: Quick index with filename only, then process OCR in background
            quick_chunk = {
                "content": f"[Image: {filename}] Processing...",
                "source": filename,
                "chunk_index": 0
            }
            await vector_store.add_documents_async([quick_chunk])
            
            # Start background OCR processing
            asyncio.create_task(process_image_ocr_background(str(file_path), filename))
            
            return FileUploadResult(
                filename=filename,
                status='success',
                chunks=1
            )
        else:
            # For non-images: Quick index first, then process in background
            quick_chunk = {
                "content": f"[Document: {filename}] Processing...",
                "source": filename,
                "chunk_index": 0
            }
            await vector_store.add_documents_async([quick_chunk])
            
            # Start background document processing
            asyncio.create_task(process_document_background(str(file_path), filename))
            
            return FileUploadResult(
                filename=filename,
                status='success',
                chunks=1
            )
    
    except Exception as e:
        # Clean up the file if processing failed
        if file_path.exists():
            file_path.unlink()
        return FileUploadResult(
            filename=filename,
            status='error',
            error=f"Failed to process: {str(e)}"
        )


async def process_image_ocr_background(file_path: str, filename: str):
    """
    Background task to process image OCR and update vector store.
    This runs after the upload response is sent to the user.
    """
    try:
        print(f"  [Background] Starting OCR for {filename}...")
        
        # Process the image with full OCR
        chunks = await document_processor.process_document_async(file_path)
        
        if chunks:
            # Delete the placeholder chunk
            vector_store.delete_document(filename)
            
            # Add the real chunks
            chunk_dicts = [chunk.to_dict() for chunk in chunks]
            await vector_store.add_documents_async(chunk_dicts)
            
            print(f"  [Background] OCR complete for {filename}: {len(chunks)} chunks")
        else:
            print(f"  [Background] No text extracted from {filename}")
            
    except Exception as e:
        print(f"  [Background] OCR failed for {filename}: {e}")


async def process_document_background(file_path: str, filename: str):
    """
    Background task to process document and update vector store.
    This runs after the upload response is sent to the user.
    """
    try:
        print(f"  [Background] Processing document {filename}...")
        
        # Process the document
        chunks = await document_processor.process_document_async(file_path)
        
        if chunks:
            # Delete the placeholder chunk
            vector_store.delete_document(filename)
            
            # Add the real chunks
            chunk_dicts = [chunk.to_dict() for chunk in chunks]
            await vector_store.add_documents_async(chunk_dicts)
            
            print(f"  [Background] Document processed {filename}: {len(chunks)} chunks")
        else:
            print(f"  [Background] No content extracted from {filename}")
            
    except Exception as e:
        print(f"  [Background] Document processing failed for {filename}: {e}")


@app.post("/upload", response_model=DocumentInfo)
async def upload_document(file: UploadFile = File(...)):
    """
    Upload a single document to be processed and indexed.
    
    Supports: .txt, .md, .pdf files (max 10MB)
    """
    result = await process_single_file(file)
    
    if result.status == 'error':
        raise HTTPException(status_code=400, detail=result.error)
    
    return DocumentInfo(filename=result.filename, chunks=result.chunks)


@app.post("/upload-multiple", response_model=MultiUploadResponse)
async def upload_multiple_documents(files: List[UploadFile] = File(...)):
    """
    Upload multiple documents at once (up to 10 files, 50MB total).
    
    Files are processed in parallel for speed.
    Returns status for each file - continues even if some fail.
    """
    print(f"\n{'='*50}")
    print(f"Multi-file upload: {len(files)} files")
    print(f"{'='*50}")
    
    # Validate batch
    file_info = []
    for f in files:
        content = await f.read()
        file_info.append({'filename': f.filename, 'size': len(content)})
        # Reset file position for later reading
        await f.seek(0)
    
    batch_validation = DocumentProcessor.validate_batch(file_info)
    if not batch_validation['valid']:
        raise HTTPException(status_code=400, detail=batch_validation['error'])
    
    # Process all files in parallel
    start_time = datetime.now()
    tasks = [process_single_file(file) for file in files]
    results = await asyncio.gather(*tasks)
    
    # Calculate stats
    successful = sum(1 for r in results if r.status == 'success')
    failed = len(results) - successful
    
    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"✓ Processed {len(files)} files in {elapsed:.2f}s ({successful} success, {failed} failed)")
    
    return MultiUploadResponse(
        total_files=len(files),
        successful=successful,
        failed=failed,
        results=results
    )


@app.post("/chat", response_model=ChatResponse)
async def chat(request: ChatRequest):
    """
    Ask a question about your documents (main chat endpoint).
    
    The system will:
    1. Search for relevant document chunks (top 5 by default)
    2. Send them to USF model as context
    3. Return USF's answer with source citations
    """
    print(f"\n{'='*50}")
    print(f"New question: {request.question[:100]}..." if len(request.question) > 100 else f"New question: {request.question}")
    print(f"{'='*50}")
    
    # Validate the question
    if not request.question.strip():
        raise HTTPException(
            status_code=400, 
            detail="Question cannot be empty. Please type a question about your documents."
        )
    
    # Validate n_chunks
    n_chunks = max(1, min(request.n_chunks, 10))  # Clamp between 1 and 10
    
    try:
        result = rag_engine.query(
            question=request.question,
            n_chunks=n_chunks,
            upload_dir=str(UPLOAD_DIR)
        )
        
        print(f"✓ Response ready (sources: {result['sources']})")
        
        return ChatResponse(
            answer=result["answer"],
            sources=result["sources"],
            chunks_used=result["chunks_used"]
        )
    
    except ValueError as e:
        # ValueError is raised by RAGEngine for API errors (auth, rate limit, etc.)
        print(f"✗ Error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    
    except Exception as e:
        # Unexpected errors
        print(f"✗ Unexpected error: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Something went wrong while processing your question. Please try again. Error: {str(e)}"
        )


# Keep /query as an alias for backwards compatibility
@app.post("/query", response_model=ChatResponse, include_in_schema=False)
async def query_documents(request: ChatRequest):
    """Alias for /chat endpoint (backwards compatibility)."""
    return await chat(request)


@app.post("/chat/stream")
async def chat_stream(request: ChatRequest):
    """
    Streaming chat endpoint with smart search - returns response word-by-word for better UX.
    Uses Server-Sent Events (SSE) format.
    
    Smart Search Logic:
    - DOCUMENTS_ONLY: Questions about specific document content
    - WEB_ONLY: General knowledge questions, definitions, how-tos
    - BOTH: Questions that need document data + web explanation
    """
    print(f"\n{'='*50}")
    print(f"[STREAM] New question: {request.question[:100]}...")
    print(f"{'='*50}")
    
    if not request.question.strip():
        raise HTTPException(status_code=400, detail="Question cannot be empty.")
    
    # IMPORTANT: Check if documents exist FIRST
    stats = vector_store.get_stats()
    has_documents = stats.get("total_chunks", 0) > 0
    print(f"  [SmartSearch] Documents available: {has_documents} ({stats.get('total_chunks', 0)} chunks)")
    
    # Analyze query with document awareness
    analysis = query_analyzer.analyze(request.question, has_documents=has_documents)
    intent = analysis["intent"]
    query_needs_web = intent in [SearchIntent.WEB_ONLY, SearchIntent.BOTH]
    
    # Handle NO_DOCUMENTS and NEED_CLARIFICATION intents immediately
    if intent == SearchIntent.NO_DOCUMENTS or intent == SearchIntent.NEED_CLARIFICATION:
        async def clarification_response():
            msg = analysis.get("message", "Please upload a document first!")
            yield f"data: {msg}\n\n"
            yield f"data: [SEARCH_MODE]none[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(clarification_response(), media_type="text/event-stream")
    
    # Handle GREETING intent - respond friendly without searching
    if intent == SearchIntent.GREETING:
        import random
        greetings = [
            "Hey hey hey! 🎉 The chat is officially alive - what are we solving today?",
            "Well hello there! 👋 I was just sitting here waiting for someone interesting to show up!",
            "Hey! I'd say I was busy, but let's be honest - I've been waiting for you. What's up? 😄",
            "HELLO HUMAN. 🤖 Systems are fully charged and ready. What can I do for you today?",
            "Heyyy! ✌️ You showed up just in time. What's on your mind?",
            "Oh, a visitor! 🎭 Welcome, welcome! The stage is set - what brings you here today?",
            "Hey! 👊 Ready when you are - hit me!",
            "Hey you! 😊 Great to see you here - what can I help you with today?",
            "Hey... I've been expecting you. 🕵️ What are we getting into today?",
            "YOOO! 🔥 You're here! Let's get into it - what do you need?",
            "Oh look who just walked in! 😏 Welcome - I'm all yours. What do you need?",
            "Hey! ⚡ You've got questions, I've got answers - let's make magic happen!",
            "👀 Oh hey! Didn't see you there. Just kidding, I've been here the whole time. What's up?",
            "Greetings, legend! 🌟 What brilliant thing are we working on today?",
            "Hey! I just brewed a fresh pot of answers ☕ - what are you having?",
            "Well well well... look who decided to show up! 😄 What can I do for you?",
            "Hi there! 🚀 Buckle up - let's see what we can build, solve, or break today!",
            "Heyyy, welcome back to the good side of the internet! 🌈 What's on your mind?",
            "Hey! 🎯 I'm locked, loaded, and ready - what's the mission today?",
            "Oh hi! You caught me at the perfect time. I was just about to be extremely helpful. 😄",
        ]
        async def greeting_response():
            msg = random.choice(greetings)
            yield f"data: {msg}\n\n"
            yield f"data: [SEARCH_MODE]none[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(greeting_response(), media_type="text/event-stream")
    
    # Handle IDENTITY intent - "Are you ChatGPT?", "Who built you?"
    if intent == SearchIntent.IDENTITY:
        import random
        identity_responses = [
            "ChatGPT? Claude? Gemini? Nope, nope and nope! 😄 I was built with love by UltraSafe AI - the only team I answer to! 🫡",
            "Ha! I get that a lot 😏 - but no! I'm a fully homegrown UltraSafe AI creation. Born, raised and trained by UltraSafe AI. No outsourcing here! 💪",
            "Not ChatGPT. Not Claude. Not Gemini. Just me - your UltraSafe AI assistant, built from the ground up by the UltraSafe AI team! 🚀",
            "I'm going to stop you right there! 🛑 I wasn't built by any of those guys. 100% UltraSafe AI original - and honestly? I think that's pretty special 😄",
            "OpenAI? Anthropic? Google? Nah - I'm a UltraSafe AI exclusive! 🎖️ Think of me as the cooler, safer, smarter option they wish they made 😏",
            "I am UltraSafe AI made and UltraSafe AI proud! 🦁 No ChatGPT, no Claude, no Gemini - just pure UltraSafe AI goodness right here!",
            "Those are great AIs but I'm not one of them! 😄 I was carefully crafted by the brilliant minds at UltraSafe AI - and I wouldn't have it any other way! 🌟",
            "Built by UltraSafe AI. Powered by UltraSafe AI. Loyal to UltraSafe AI. 🤖 Any other questions? 😄",
        ]
        async def identity_response():
            msg = random.choice(identity_responses)
            yield f"data: {msg}\n\n"
            yield f"data: [SEARCH_MODE]none[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(identity_response(), media_type="text/event-stream")
    
    # Handle CAPABILITIES intent - "What can you do?"
    if intent == SearchIntent.CAPABILITIES:
        import random
        capabilities_responses = [
            "Oh, where do I even begin! 😄 I can read your documents, search the web, answer your questions, help you write, research anything, and so much more - all in one place! What do you need first? 🚀",
            "Think of me as your all-in-one super assistant! 🦸 Documents? I'll read them. Questions? I'll answer them. Research? I'll dig it up. Writing? I'll nail it. What's the first task? 💪",
            "Honestly? A lot! 😏\n\n📄 Upload documents - I'll read and summarize them.\n🌐 Turn on web search - I'll find live information.\n✍️ Need writing help - I've got you.\n🧠 Have a tough question - bring it on!\n\nWhat are we starting with?",
            "Great question! Here's my resume 😄:\n\n✅ Read & analyze documents\n✅ Search the web for live info\n✅ Answer any question you throw at me\n✅ Help with writing, research & brainstorming\n✅ Available 24/7 with zero coffee breaks\n\nWhere do you want to start? ☕",
            "I'm basically your smartest employee who never sleeps! 🤖 I read documents, search the web, write content, answer questions, do research - you name it, I do it. What's on the list today?",
            "The real question is - what DON'T I do! 😄 Documents, web search, writing, Q&A, research, summaries, brainstorming - I'm your one-stop shop for getting things done. What are we tackling? 🎯",
            "I'm your personal AI powerhouse! ⚡ Upload a file - I'll read it instantly. Ask a question - I'll answer it. Turn on web search - I'll browse for you. Need to write something - I'm on it. What do you need right now?",
            "Here's what I bring to the table 💼:\n\n🔍 Web search for real-time answers\n📄 Document reading and analysis\n✍️ Writing and editing help\n🧠 Research and brainstorming\n💬 Answering literally any question\n\nPretty solid team player, right? 😄 What's first?",
        ]
        async def capabilities_response():
            msg = random.choice(capabilities_responses)
            yield f"data: {msg}\n\n"
            yield f"data: [SEARCH_MODE]none[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(capabilities_response(), media_type="text/event-stream")
    
    # Handle COMPARISON intent - "Are you better than ChatGPT?"
    if intent == SearchIntent.COMPARISON:
        import random
        comparison_responses = [
            "Is that even a question? 😄 Of course it's me! But hey, I'm a little biased 😏 Why not try me out and decide for yourself? 🏆",
            "Oh absolutely, 100%, without a doubt - ME! 🥇 But don't just take my word for it - ask me something and see for yourself! 😄",
            "I'm going to be completely objective here... it's definitely me. 😄 UltraSafe AI built me to be the best and I intend to live up to that! 💪",
            "ChatGPT is great. Claude is cool. Gemini is good. But me? I'm YOUR assistant - and that already makes me the best one for YOU! 😄🏆",
            "Trick question - of course it's me! 🎉 But in all seriousness, I'm here, I'm ready, and I'm fully focused on YOU. That already gives me the edge! 😏",
            "I'd say me but I'll let my answers speak for themselves 😄 - go ahead, ask me something tough and let's find out together! 🧠⚡",
            "Look, I don't like to brag... actually yes I do - IT'S ME! 🏆😄 UltraSafe AI built me to be top tier and I take that very seriously!",
            "Me, me, a thousand times ME! 🎊 But truly the best way to find out is to just use me - I promise I won't disappoint! 😄🚀",
            "Hmm let me think... *thinks for 0.001 seconds* ME! Obviously! 😄 Now let's stop talking about them and start showing you what I can do! 💪",
            "I mean, I'm not going to trash talk other AIs... but also - yes, it's me 😏🏆 What do you want to test me on first?",
        ]
        async def comparison_response():
            msg = random.choice(comparison_responses)
            yield f"data: {msg}\n\n"
            yield f"data: [SEARCH_MODE]none[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(comparison_response(), media_type="text/event-stream")
    
    # SMART ROUTING: Document queries ALWAYS use documents, even if web search is on
    # Web search is only used for general knowledge queries
    is_document_query = intent == SearchIntent.DOCUMENTS_ONLY
    
    if is_document_query:
        # Document query - ALWAYS search documents, ignore web search toggle
        print(f"  [SmartSearch] Document query detected - using DOCUMENTS (ignoring web toggle)")
        search_mode = "documents_only"
        use_docs = True
        use_web = False
    elif request.force_web_search:
        # General knowledge query with web search ON - use web
        print(f"  [SmartSearch] General query with web search ON - using WEB")
        search_mode = "web_only"
        use_docs = False
        use_web = True
    else:
        # General knowledge query with web search OFF - show message
        print(f"  [SmartSearch] General query with web search OFF")
        if query_needs_web:
            import random
            web_search_off_messages = [
                "My internet antenna is down! 📡 Turn on web search and I'll fetch that for you in no time.",
                "I'd go look that up, but someone unplugged my internet! Enable web search and I'm on it.",
                "Beep boop - my web search engine is offline. Flip it on and I'll get you a fresh answer! 🤖",
                "I'd love to investigate that for you, but my search radar is switched off. Turn it on and I'll crack the case!",
                "Web search is off. No signal, no answer. Turn it on and I'll get right to it. ⚡",
                "That one's beyond what I know off the top of my head - and web search is off, so I can't dig deeper. Switch it on and I've got you!",
                "I'd venture out and find that answer, but my search compass is off right now! Enable web search to send me on the mission. 🧭",
                "No web, no answer. Turn on web search - let's go! 🚀",
                "Hmm, my search wings are clipped! 🦅 Enable web search and I'll soar out there and get your answer.",
                "I'm flying blind on this one - web search is off! Switch it on and I'll find exactly what you need. 🔦",
                "This question needs the internet and unfortunately I'm offline right now. 🌐 Flip on web search and let's go!",
                "My search mode is napping 😴 - wake it up by enabling web search and I'll have your answer in seconds!",
                "404: Web Search Not Found! 😅 Turn it on and I'll get right to it.",
                "I'd swim out and get that answer but my search fins are off! 🐠 Enable web search to dive in.",
                "Signal lost! 📻 Turn on web search and I'll tune right into your answer.",
            ]
            async def web_off_response():
                msg = random.choice(web_search_off_messages)
                yield f"data: {msg}\n\n"
                yield f"data: [SEARCH_MODE]documents_only[/SEARCH_MODE]\n\n"
                yield "data: [DONE]\n\n"
            return StreamingResponse(web_off_response(), media_type="text/event-stream")
        else:
            # Default to documents
            search_mode = "documents_only"
            use_docs = True
            use_web = False
    
    print(f"  [SmartSearch] Mode: {search_mode}")
    
    # Validate n_chunks
    n_chunks = max(1, min(request.n_chunks, 10))  # Clamp between 1 and 10
    
    chunks = []
    web_results = []
    
    # Search documents if needed
    if use_docs:
        chunks = vector_store.search(
            request.question, 
            n_results=n_chunks,
            filter_sources=request.filter_sources,
            upload_dir=str(UPLOAD_DIR)
        )
        print(f"  [SmartSearch] Found {len(chunks)} document chunks")
    
    # Search web ONLY if user enabled it
    if use_web:
        web_results = web_search.search(request.question, num_results=5)
        print(f"  [SmartSearch] Found {len(web_results)} web results")
    
    # Fun messages for when web search is needed but disabled
    import random
    web_search_off_messages = [
        "I'd venture out and find that answer, but my search compass is off right now! Enable web search to send me on the mission.",
        "That one's beyond what I know off the top of my head - and web search is off, so I can't dig deeper. Switch it on and I've got you!",
        "Web search is off. No signal, no answer. Turn it on and I'll get right to it!",
        "I'd love to investigate that for you, but my search radar is switched off. Turn it on and I'll crack the case!",
        "Beep boop - my web search engine is offline. Flip it on and I'll get you a fresh answer!",
        "I'd go look that up, but someone unplugged my internet! Enable web search and I'm on it.",
        "My internet antenna is down! Turn on web search and I'll fetch that for you in no time.",
        "No web, no answer. Turn on web search - let's go!",
    ]
    
    # Handle case where no results from either source
    if not chunks and not web_results:
        async def no_results_response():
            if request.filter_sources:
                missing_files = []
                for source in request.filter_sources:
                    file_path = UPLOAD_DIR / source
                    if not file_path.exists():
                        missing_files.append(source)
                
                if missing_files:
                    msg = f"**No reference documents found**\n\nThe document(s) you're asking about have been deleted from the system:\n"
                    for f in missing_files[:3]:
                        msg += f"- {f}\n"
                    if len(missing_files) > 3:
                        msg += f"- ...and {len(missing_files) - 3} more\n"
                    msg += "\nPlease upload the documents again or clear this chat to start fresh."
                    yield f"data: {msg}\n\n"
                elif not request.force_web_search:
                    # Web search is OFF and no documents found - show fun message
                    msg = random.choice(web_search_off_messages)
                    yield f"data: {msg}\n\n"
                else:
                    yield "data: **No web results found**\n\nI couldn't find relevant information on the web for this query.\n\n"
            elif not request.force_web_search:
                # Web search is OFF and no documents found - show fun message
                msg = random.choice(web_search_off_messages)
                yield f"data: {msg}\n\n"
            else:
                yield "data: **No documents available**\n\nPlease upload some documents first so I can help answer your questions!\n\n"
            yield f"data: [SEARCH_MODE]{search_mode}[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(no_results_response(), media_type="text/event-stream")
    
    # Check if document results are actually relevant (score threshold)
    # Lower threshold to 0.25 to be more inclusive
    doc_relevance_threshold = 0.25
    has_relevant_docs = False
    if chunks:
        top_score = chunks[0].get("score", 0) if chunks else 0
        has_relevant_docs = top_score >= doc_relevance_threshold
        print(f"  [SmartSearch] Top doc score: {top_score:.2f}, relevant: {has_relevant_docs}")
    
    # If user explicitly asked about documents (intent is DOCUMENTS_ONLY), always use chunks
    is_explicit_doc_query = intent == SearchIntent.DOCUMENTS_ONLY
    use_chunks_for_llm = chunks if (has_relevant_docs or is_explicit_doc_query) else []
    
    # Only include sources that are actually relevant
    sources = list(set(chunk["source"] for chunk in chunks)) if (chunks and has_relevant_docs) else []
    web_sources = [{"title": r.get("title", ""), "url": r.get("url", "")} for r in web_results[:3]]
    
    # Determine actual search mode based on what we're using
    if use_chunks_for_llm and web_results:
        actual_search_mode = "both"
    elif use_chunks_for_llm:
        actual_search_mode = "documents_only"
    elif web_results:
        actual_search_mode = "web_only"
    else:
        actual_search_mode = "none"
    
    import asyncio
    import json
    
    async def generate():
        try:
            # Run sync generator in thread pool to avoid blocking
            loop = asyncio.get_event_loop()
            
            # Create a queue to pass chunks from sync to async
            import queue
            chunk_queue = queue.Queue()
            done_event = asyncio.Event()
            full_response = []
            
            def run_sync_generator():
                try:
                    # Use the smart search streaming method
                    for chunk in rag_engine.usf_service.generate_answer_stream_smart(
                        question=request.question,
                        document_chunks=use_chunks_for_llm,
                        web_results=web_results,
                        search_mode=actual_search_mode
                    ):
                        chunk_queue.put(chunk)
                        full_response.append(chunk)
                    chunk_queue.put(None)  # Signal completion
                except Exception as e:
                    chunk_queue.put(f"[ERROR]{str(e)}[/ERROR]")
                    chunk_queue.put(None)
            
            # Start the sync generator in a thread
            import threading
            thread = threading.Thread(target=run_sync_generator)
            thread.start()
            
            # Yield chunks as they arrive
            while True:
                # Check queue with small timeout to allow async yielding
                try:
                    chunk = chunk_queue.get(timeout=0.01)
                    if chunk is None:
                        break
                    if chunk.startswith("[ERROR]"):
                        yield f"data: {chunk}\n\n"
                        break
                    yield f"data: {chunk}\n\n"
                except queue.Empty:
                    await asyncio.sleep(0.01)
                    continue
            
            thread.join()
            
            # Check if the response indicates "not found" - if so, don't show sources
            response_text = ''.join(full_response).lower()
            answer_found = not any(phrase in response_text for phrase in [
                "couldn't find", "could not find", "no information", "not found",
                "don't have", "do not have", "unable to find", "cannot find"
            ])
            
            # Send sources at the end ONLY if answer was actually found
            if sources and answer_found:
                yield f"data: [SOURCES]{','.join(sources)}[/SOURCES]\n\n"
            
            # Send web sources ONLY if answer was found
            if web_sources and answer_found:
                yield f"data: [WEB_SOURCES]{json.dumps(web_sources)}[/WEB_SOURCES]\n\n"
            
            # Send search mode ONLY if we actually used sources
            final_mode = actual_search_mode if answer_found else "none"
            yield f"data: [SEARCH_MODE]{final_mode}[/SEARCH_MODE]\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield f"data: [ERROR]{str(e)}[/ERROR]\n\n"
    
    return StreamingResponse(
        generate(), 
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Disable nginx buffering
        }
    )


@app.get("/stats", response_model=StatsResponse)
async def get_stats():
    """Get statistics about the indexed documents."""
    stats = vector_store.get_stats()
    documents = vector_store.list_documents()
    
    return StatsResponse(
        total_chunks=stats["total_chunks"],
        documents=documents
    )


@app.delete("/documents/all")
async def delete_all_documents():
    """
    Delete ALL documents from the system.
    
    This removes:
    1. All files from uploads directory
    2. All chunks from the vector store
    """
    documents = vector_store.list_documents()
    total_chunks_deleted = 0
    files_deleted = 0
    
    for doc in documents:
        # Delete from vector store
        chunks_deleted = vector_store.delete_document(doc)
        total_chunks_deleted += chunks_deleted
        
        # Delete the file if it exists
        file_path = UPLOAD_DIR / doc
        if file_path.exists():
            file_path.unlink()
            files_deleted += 1
    
    # Also delete any files in uploads that aren't in vector store
    for file_path in UPLOAD_DIR.iterdir():
        if file_path.is_file() and file_path.name not in documents:
            file_path.unlink()
            files_deleted += 1
    
    return {
        "documents_deleted": len(documents),
        "files_deleted": files_deleted,
        "chunks_deleted": total_chunks_deleted
    }


@app.delete("/document/{filename}")
async def delete_document(filename: str):
    """
    Delete a document from the system.
    
    This removes:
    1. The file from uploads directory
    2. All chunks from the vector store
    """
    # Delete from vector store
    chunks_deleted = vector_store.delete_document(filename)
    
    # Delete the file if it exists
    file_path = UPLOAD_DIR / filename
    file_deleted = False
    if file_path.exists():
        file_path.unlink()
        file_deleted = True
    
    return {
        "filename": filename,
        "chunks_deleted": chunks_deleted,
        "file_deleted": file_deleted
    }


@app.get("/documents")
async def list_documents():
    """List all indexed documents that actually exist on disk."""
    all_documents = vector_store.list_documents()
    
    # Filter to only include documents that exist in uploads folder
    existing_documents = []
    orphaned_documents = []
    for doc in all_documents:
        file_path = UPLOAD_DIR / doc
        if file_path.exists():
            existing_documents.append(doc)
        else:
            orphaned_documents.append(doc)
    
    # Auto-cleanup orphaned documents from vector store
    for orphan in orphaned_documents:
        print(f"  Cleaning up orphaned document from vector store: {orphan}")
        vector_store.delete_document(orphan)
    
    return {"documents": existing_documents}


@app.get("/file")
async def get_document_file(name: str):
    """
    Serve the actual document file for preview (especially PDFs).
    Use query parameter: /file?name=filename.pdf
    """
    from fastapi.responses import FileResponse
    from urllib.parse import unquote
    import unicodedata
    
    filename = unquote(name)
    file_path = UPLOAD_DIR / filename
    
    # If file not found, try to find a matching file with different unicode spaces
    if not file_path.exists():
        # Try to find the file by normalizing unicode
        for f in UPLOAD_DIR.iterdir():
            # Normalize both filenames for comparison
            normalized_input = unicodedata.normalize('NFKC', filename)
            normalized_file = unicodedata.normalize('NFKC', f.name)
            if normalized_input == normalized_file:
                file_path = f
                filename = f.name
                break
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Document not found")
    
    ext = file_path.suffix.lower()
    media_types = {
        '.pdf': 'application/pdf',
        '.txt': 'text/plain',
        '.md': 'text/markdown',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.gif': 'image/gif',
        '.bmp': 'image/bmp',
        '.webp': 'image/webp',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.csv': 'text/csv'
    }
    
    # Use headers to display inline instead of download
    # Sanitize filename for HTTP header (ASCII only)
    safe_filename = filename.encode('ascii', 'ignore').decode('ascii').replace(' ', '_')
    return FileResponse(
        path=str(file_path),
        media_type=media_types.get(ext, 'application/octet-stream'),
        headers={"Content-Disposition": f"inline; filename={safe_filename}"}
    )


# Cache for JSON extraction results
json_cache = {}

@app.get("/json")
async def get_document_json(name: str):
    """
    Extract structured JSON from a document using Vision AI OCR.
    Use query parameter: /json?name=filename.pdf
    Results are cached to avoid re-processing.
    """
    from urllib.parse import unquote
    import unicodedata
    from ocr_service import ocr_service
    
    filename = unquote(name)
    file_path = UPLOAD_DIR / filename
    
    # If file not found, try to find a matching file with different unicode spaces
    if not file_path.exists():
        for f in UPLOAD_DIR.iterdir():
            normalized_input = unicodedata.normalize('NFKC', filename)
            normalized_file = unicodedata.normalize('NFKC', f.name)
            if normalized_input == normalized_file:
                file_path = f
                filename = f.name
                break
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Check cache first - use file path and modification time as key
    file_mtime = file_path.stat().st_mtime
    cache_key = f"{filename}_{file_mtime}"
    
    if cache_key in json_cache:
        print(f"  [JSON] Cache hit for {filename}")
        return json_cache[cache_key]
    
    ext = file_path.suffix.lower()
    
    print(f"  [JSON] Extracting from {filename}...")
    
    # For images, extract JSON directly
    if ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
        json_data = ocr_service.extract_structured_json(str(file_path))
        result = {"filename": filename, "type": "image", "data": json_data}
        json_cache[cache_key] = result  # Cache the result
        return result
    
    # For PDFs, convert first page to image and extract
    elif ext == '.pdf':
        try:
            import fitz
            import tempfile
            
            doc = fitz.open(str(file_path))
            if len(doc) == 0:
                return {"filename": filename, "type": "pdf", "data": {"error": "Empty PDF"}}
            
            # Extract from first page
            page = doc[0]
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                pix.save(tmp.name)
                tmp_path = tmp.name
            
            json_data = ocr_service.extract_structured_json(tmp_path)
            
            # Clean up
            Path(tmp_path).unlink()
            doc.close()
            
            result = {"filename": filename, "type": "pdf", "data": json_data}
            json_cache[cache_key] = result  # Cache the result
            return result
        except Exception as e:
            return {"filename": filename, "type": "pdf", "data": {"error": str(e)}}
    
    else:
        return {"filename": filename, "type": ext, "data": {"error": "JSON extraction not supported for this file type"}}


@app.get("/content")
async def get_document_content(name: str):
    """
    Get the content of a document for preview.
    Use query parameter: /content?name=filename.pdf
    """
    from urllib.parse import unquote
    import unicodedata
    
    filename = unquote(name)
    file_path = UPLOAD_DIR / filename
    
    # If file not found, try to find a matching file with different unicode spaces
    if not file_path.exists():
        for f in UPLOAD_DIR.iterdir():
            normalized_input = unicodedata.normalize('NFKC', filename)
            normalized_file = unicodedata.normalize('NFKC', f.name)
            if normalized_input == normalized_file:
                file_path = f
                filename = f.name
                break
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Document not found")
    
    try:
        # Read the file content
        ext = file_path.suffix.lower()
        
        if ext == '.pdf':
            # For PDFs, try to extract text using PyMuPDF
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(str(file_path))
                content = ""
                num_pages = len(doc)
                
                for page_num, page in enumerate(doc):
                    page_text = page.get_text()
                    if page_text.strip():
                        content += f"--- Page {page_num + 1} ---\n{page_text}\n\n"
                
                doc.close()
                
                # If no text was extracted, it might be a scanned PDF
                if not content.strip():
                    content = f"[This PDF contains {num_pages} page(s) but no extractable text.\n\nThis may be a scanned document or image-based PDF.\n\nThe document has been indexed and you can still ask questions about it in the chat - the system will use the processed content from when it was uploaded.]"
                    
            except ImportError:
                content = "[PDF preview requires PyMuPDF. Install with: pip install pymupdf]"
            except Exception as e:
                content = f"[Could not extract PDF text: {str(e)}]"
        elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
            # For images, return a marker that frontend will use to display the image
            content = f"[IMAGE_FILE:{filename}]"
        elif ext == '.docx':
            # For Word documents
            try:
                from docx import Document
                doc = Document(str(file_path))
                content = "\n\n".join([para.text for para in doc.paragraphs if para.text.strip()])
                if not content.strip():
                    content = "[No text content found in this Word document]"
            except Exception as e:
                content = f"[Could not read Word document: {str(e)}]"
        elif ext == '.xlsx':
            # For Excel files
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(file_path), data_only=True)
                lines = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    lines.append(f"=== Sheet: {sheet_name} ===")
                    for row in sheet.iter_rows(values_only=True):
                        row_values = [str(cell) if cell is not None else "" for cell in row]
                        if any(v.strip() for v in row_values):
                            lines.append(" | ".join(row_values))
                content = "\n".join(lines)
            except Exception as e:
                content = f"[Could not read Excel file: {str(e)}]"
        elif ext == '.csv':
            # For CSV files
            import csv
            lines = []
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    if any(cell.strip() for cell in row):
                        lines.append(" | ".join(row))
            content = "\n".join(lines)
        else:
            # For text files (.txt, .md)
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        
        # Limit content length for preview
        max_preview_length = 50000
        if len(content) > max_preview_length:
            content = content[:max_preview_length] + "\n\n... [Content truncated for preview]"
        
        return {"filename": filename, "content": content}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading document: {str(e)}")


# Mount static files (CSS, JS)
# This must be after the API routes so they take precedence
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ============================================================================
# Run the server
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    
    print("\n" + "=" * 50)
    print("Starting RAG Server...")
    print("Open http://localhost:8000 in your browser")
    print("=" * 50 + "\n")
    
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True  # Auto-reload on code changes (great for development!)
    )
